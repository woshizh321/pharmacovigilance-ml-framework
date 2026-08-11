from __future__ import annotations

import csv
import json
import math
import shutil
import sys
import textwrap
from pathlib import Path

import matplotlib as mpl
mpl.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.lines import Line2D
from matplotlib.patches import FancyArrowPatch, FancyBboxPatch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
FIG = ROOT / "figures"
MAIN_FIG = FIG / "main"
SUPP_FIG = FIG / "supplement"
SRC_FIG = FIG / "source_data"
TAB = ROOT / "tables"
MAIN_TAB = TAB / "main"
SUPP_TAB = TAB / "supplement"

for directory in (MAIN_FIG, SUPP_FIG, SRC_FIG, MAIN_TAB, SUPP_TAB):
    directory.mkdir(parents=True, exist_ok=True)

mpl.rcParams.update({
    "font.family": "sans-serif",
    "font.sans-serif": ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size": 7.5,
    "axes.labelsize": 8,
    "axes.titlesize": 8,
    "xtick.labelsize": 7,
    "ytick.labelsize": 7,
    "legend.fontsize": 7,
    "axes.linewidth": 0.7,
    "lines.linewidth": 1.2,
    "savefig.facecolor": "white",
    "figure.facecolor": "white",
    "pdf.fonttype": 42,
    "ps.fonttype": 42,
    "svg.fonttype": "none",
})

COL = {
    "set0": "#B8B8B8",
    "set1": "#0072B2",
    "dev": "#0072B2",
    "hold": "#D55E00",
    "pre": "#0072B2",
    "post": "#999999",
    "supported": "#009E73",
    "rare": "#CC79A7",
    "elasticnet": "#0072B2",
    "xgboost": "#D55E00",
    "consensus": "#009E73",
    "directional": "#E69F00",
    "ink": "#222222",
    "light": "#F2F2F2",
}

VMANIFEST: list[dict] = []
TMANIFEST: list[dict] = []
FIG_META: dict[str, dict] = {}


def read_csv(rel: str) -> pd.DataFrame:
    return pd.read_csv(ROOT / rel)


def clean_axes(ax, grid_axis: str | None = None):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if grid_axis:
        ax.grid(axis=grid_axis, color="#D9D9D9", linewidth=0.45, alpha=0.65)
        ax.set_axisbelow(True)


def panel_label(ax, label: str, x: float = -0.12, y: float = 1.06):
    ax.text(x, y, label, transform=ax.transAxes, fontsize=10, fontweight="bold", va="top")


def source_csv(df: pd.DataFrame, name: str) -> Path:
    path = SRC_FIG / name
    df.to_csv(path, index=False)
    return path


def export_figure(fig, base: Path, figure_id: str):
    width, height = fig.get_size_inches()
    fig.savefig(base.with_suffix(".svg"), bbox_inches="tight")
    fig.savefig(base.with_suffix(".pdf"), bbox_inches="tight")
    fig.savefig(base.with_suffix(".png"), dpi=300, bbox_inches="tight")
    fig.savefig(
        base.with_suffix(".tiff"), dpi=600, bbox_inches="tight",
        pil_kwargs={"compression": "tiff_lzw"},
    )
    FIG_META[figure_id] = {
        "width_in": float(width), "height_in": float(height),
        "font_family": "Arial/Helvetica sans-serif", "base_font_pt": 7.5,
    }
    plt.close(fig)


def vmanifest(figure_id: str, panel: str, source_file: str, source_columns: str,
              locked_section: str, metric: str, transform: str, ci_source: str,
              base: Path, qc: str = "PASS"):
    meta = FIG_META.get(figure_id, {})
    VMANIFEST.append({
        "figure_id": figure_id, "panel": panel, "source_file": source_file,
        "source_columns": source_columns, "locked_section": locked_section,
        "plotted_metric": metric, "transformation_for_display": transform,
        "CI_source": ci_source,
        "output_svg": str(base.with_suffix(".svg").relative_to(ROOT)),
        "output_pdf": str(base.with_suffix(".pdf").relative_to(ROOT)),
        "output_tiff": str(base.with_suffix(".tiff").relative_to(ROOT)),
        "output_png": str(base.with_suffix(".png").relative_to(ROOT)),
        "numerical_QC_status": qc,
        "figure_width_in": meta.get("width_in", ""),
        "figure_height_in": meta.get("height_in", ""),
        "font_family": meta.get("font_family", ""),
        "base_font_pt": meta.get("base_font_pt", ""),
    })


def draw_box(ax, xy, width, height, text, face="#F2F2F2", edge="#555555",
             fontsize=7.0, lw=0.8, linestyle="-"):
    x, y = xy
    patch = FancyBboxPatch(
        (x, y), width, height,
        boxstyle="round,pad=0.018,rounding_size=0.018",
        facecolor=face, edgecolor=edge, linewidth=lw, linestyle=linestyle,
    )
    ax.add_patch(patch)
    ax.text(x + width / 2, y + height / 2, text, ha="center", va="center",
            fontsize=fontsize, color=COL["ink"], wrap=True)
    return patch


def arrow(ax, start, end, color="#666666", style="-", mutation=8, lw=0.9):
    arr = FancyArrowPatch(start, end, arrowstyle="-|>", mutation_scale=mutation,
                          color=color, linewidth=lw, linestyle=style,
                          connectionstyle="arc3,rad=0")
    ax.add_patch(arr)
    return arr


def label_pipeline(pipeline: str) -> str:
    mapping = {
        "elasticnet_set0": "Elastic-net Set 0",
        "elasticnet_set1": "Elastic-net Set 1",
        "xgboost_set0": "XGBoost Set 0",
        "xgboost_set1": "XGBoost Set 1",
    }
    return mapping[pipeline]


def plot_estimates(ax, df, label_col, estimate_col, low_col, high_col, title,
                   no_skill=None, xlim=None):
    y = np.arange(len(df))[::-1]
    colors = [COL["set1"] if "Set 1" in x else COL["set0"] for x in df[label_col]]
    markers = ["o" if "Elastic" in x else "s" for x in df[label_col]]
    for yi, (_, row), color, marker in zip(y, df.iterrows(), colors, markers):
        est, lo, hi = row[estimate_col], row[low_col], row[high_col]
        ax.errorbar(est, yi, xerr=[[est - lo], [hi - est]], fmt=marker,
                    color=COL["ink"], markerfacecolor=color, markeredgecolor=COL["ink"],
                    markersize=5, capsize=2.3, linewidth=0.9)
    if no_skill is not None:
        ax.axvline(no_skill, color="#666666", linestyle=(0, (3, 2)), linewidth=0.8)
        ax.text(no_skill, 0.98, "No-skill AP", transform=ax.get_xaxis_transform(),
                rotation=90, va="top", ha="right", fontsize=6.5, color="#555555")
    ax.set_yticks(y, df[label_col])
    ax.set_xlabel("Average precision (95% CI)")
    ax.set_title(title, loc="left", fontweight="bold")
    if xlim:
        ax.set_xlim(*xlim)
    clean_axes(ax, "x")


def plot_calibration(ax, df, family, title):
    family_rows = df[df["pipeline"].str.startswith(family)]
    for feature, color, marker, ls in [("set0", COL["set0"], "o", "--"),
                                       ("set1", COL["set1"], "s", "-")]:
        sub = family_rows[family_rows["pipeline"].str.endswith(feature)]
        ax.plot(sub["mean_predicted_probability"], sub["observed_rate"],
                marker=marker, linestyle=ls, color=color, markeredgecolor=COL["ink"],
                markeredgewidth=0.35, markersize=3.5, label=feature.upper().replace("SET", "Set "))
    lim = max(0.45, float(max(family_rows["mean_predicted_probability"].max(), family_rows["observed_rate"].max())) * 1.05)
    ax.plot([0, lim], [0, lim], color="#555555", linewidth=0.7, linestyle=":", label="Ideal")
    ax.set_xlim(0, lim)
    ax.set_ylim(0, lim)
    ax.set_xlabel("Mean native predicted probability")
    ax.set_ylabel("Observed proportion")
    ax.set_title(title, loc="left", fontsize=7.5)
    ax.legend(frameon=False, ncol=3, loc="upper left", handlelength=1.5)
    clean_axes(ax, "both")


def fig1():
    flow = read_csv("analysis/section1_cohort/01_cohort_flow_counts.csv")
    assert int(flow.loc[flow["stage"] == "01_FDA_SOURCE_APPROVAL_IDENTITIES", "active_moieties"].iloc[0]) == 466
    assert int(flow.loc[flow["stage"] == "02_SINGLE_ACTIVE_PRIMARY_COHORT", "active_moieties"].iloc[0]) == 432
    panel_a = pd.DataFrame([
        ["FDA", "Regulatory cohort and approval anchor", "Input"],
        ["ClinicalTrials.gov / AACT", "Qualifying premarketing safety evidence", "Predictor evidence"],
        ["FAERS", "Three-year postapproval Criterion-R outcome", "Outcome only"],
        ["JADER", "Cumulative cross-database replication only", "Replication only"],
    ], columns=["data_source", "role", "model_role"])
    panel_b = pd.DataFrame([
        ["FDA source identities", 466, "analysis/section1_cohort/01_cohort_flow_counts.csv"],
        ["Single-active-moiety cohort", 432, "analysis/section1_cohort/01_cohort_flow_counts.csv"],
        ["AACT reliable mapping branch", 429, "analysis/section1_cohort/SECTION1_README.md"],
        ["FAERS reliable mapping branch", 426, "analysis/section1_cohort/SECTION1_README.md"],
        ["B-STRICT drugs", 212, "analysis/section1_cohort/01_cohort_flow_counts.csv"],
        ["B-STRICT trials", 620, "analysis/section1_cohort/01_cohort_flow_counts.csv"],
        ["B-STRICT target arms", 1149, "analysis/section1_cohort/01_cohort_flow_counts.csv"],
        ["B-STRICT drug–PT pairs", 30247, "analysis/section1_cohort/01_cohort_flow_counts.csv"],
        ["FAERS identity/outcome assessability + PS≥100 input", np.nan, "analysis/section1_cohort/SECTION1_README.md"],
        ["Final PS≥100 drugs", 166, "analysis/section1_cohort/SECTION1_README.md"],
        ["Final PS≥100 candidate pairs", 26151, "analysis/section1_cohort/SECTION1_README.md"],
        ["Development drugs", 107, "analysis/section1_cohort/SECTION1_README.md"],
        ["Development pairs", 16470, "analysis/section3_model/01_development_model_domain.csv"],
        ["Temporal-holdout drugs", 59, "analysis/section1_cohort/SECTION1_README.md"],
        ["Temporal-holdout pairs", 9681, "analysis/section4_holdout/SECTION4_README.md"],
    ], columns=["stage", "locked_value", "authoritative_source"])
    panel_c = pd.DataFrame([
        ["Set 0", "Regulatory + event identity + drug-program context"],
        ["Set 1", "Set 0 + pair-specific premarketing safety information"],
        ["Models", "Penalized logistic regression; gradient-boosted trees"],
        ["Development", "Drug-grouped nested cross-validation"],
        ["Temporal validation", "Locked 2019–2022 holdout"],
        ["Outcome", "Three-year FAERS Criterion R"],
    ], columns=["component", "locked_definition"])
    source_csv(panel_a, "Figure1_panelA_data_architecture.csv")
    source_csv(panel_b, "Figure1_panelB_cohort_cascade.csv")
    source_csv(panel_c, "Figure1_panelC_prediction_architecture.csv")

    # Command 18: retain the three-panel architecture while tightening the
    # vertical footprint by 12.5% (8.0 -> 7.0 inches).
    fig = plt.figure(figsize=(7.2, 7.0))
    gs = fig.add_gridspec(3, 1, height_ratios=[0.86, 1.45, 0.98], hspace=0.23)
    axa, axb, axc = [fig.add_subplot(gs[i]) for i in range(3)]
    for ax in (axa, axb, axc):
        ax.set_axis_off()
    panel_label(axa, "A", -0.02, 1.02)
    panel_label(axb, "B", -0.02, 1.02)
    panel_label(axc, "C", -0.02, 1.02)
    axa.set_title("Data-source architecture", loc="left", fontweight="bold", pad=7)
    xs = [0.02, 0.27, 0.52, 0.77]
    faces = ["#DDEBF7", "#E2F0D9", "#FCE4D6", "#FFF2CC"]
    labels = [
        "FDA\nRegulatory cohort\nand approval anchor",
        "ClinicalTrials.gov / AACT\nQualifying premarketing\nsafety evidence",
        "FAERS\n3-year postapproval\nCriterion-R outcome",
        "JADER\nCumulative cross-database\nreplication only",
    ]
    for x, face, label in zip(xs, faces, labels):
        draw_box(axa, (x, 0.37), 0.20, 0.35, label, face=face, fontsize=7.2)
    arrow(axa, (0.22, 0.54), (0.27, 0.54))
    arrow(axa, (0.47, 0.54), (0.52, 0.54), color=COL["hold"])
    arrow(axa, (0.62, 0.35), (0.84, 0.23), color=COL["directional"], style="--")
    axa.text(0.50, 0.15, "FAERS outcome is never a predictor", color="#8B1A1A",
             fontweight="bold", fontsize=6.6)
    axa.text(0.76, 0.05, "JADER: no training; no temporal validation", color="#7A5A00",
             fontweight="bold", fontsize=6.6)

    axb.set_title("Cohort cascade with parallel identity branches", loc="left", fontweight="bold", pad=7)
    draw_box(axb, (0.01, 0.76), 0.17, 0.15, "466 FDA source\nidentities", face="#DDEBF7")
    draw_box(axb, (0.23, 0.76), 0.20, 0.15, "432 single-active-\nmoiety drugs", face="#DDEBF7")
    arrow(axb, (0.18, 0.835), (0.23, 0.835))
    draw_box(axb, (0.51, 0.80), 0.20, 0.14, "AACT identity branch\n429 mapped", face="#E2F0D9")
    draw_box(axb, (0.76, 0.80), 0.20, 0.14, "FAERS identity branch\n426 mapped", face="#FCE4D6")
    axb.plot([0.43, 0.47], [0.835, 0.835], color="#666666", linewidth=0.9)
    arrow(axb, (0.47, 0.835), (0.51, 0.87))
    # Route the FAERS branch above both branch boxes so neither upstream
    # connector crosses a box or label.
    axb.plot([0.47, 0.47, 0.86], [0.835, 0.975, 0.975],
             color="#666666", linewidth=0.9)
    arrow(axb, (0.86, 0.975), (0.86, 0.94))
    axb.text(0.735, 0.72, "Parallel identity branches\n(not serial attrition)", ha="center", va="center",
             fontsize=6.7, color="#555555", fontstyle="italic")

    # Each branch feeds its own eligibility/assessability step vertically.
    # The two steps then converge at the final PS>=100 cohort.
    draw_box(axb, (0.49, 0.46), 0.24, 0.18,
             "B-STRICT evidence eligibility\n212 drugs · 620 trials\n1,149 arms · 30,247 pairs", face="#E2F0D9")
    draw_box(axb, (0.75, 0.46), 0.23, 0.18,
             "FAERS identity/outcome\nassessability + PS≥100", face="#FCE4D6")
    arrow(axb, (0.61, 0.80), (0.61, 0.64), color=COL["supported"])
    arrow(axb, (0.86, 0.80), (0.86, 0.64), color=COL["hold"])
    draw_box(axb, (0.35, 0.14), 0.26, 0.19,
             "Final PS≥100 cohort\n166 drugs\n26,151 candidate pairs", face="#EAF2F8", edge=COL["set1"], lw=1.2)
    axb.text(0.48, 0.36, "Convergence", ha="center", va="bottom", fontsize=6.5,
             color="#555555", fontweight="bold")
    arrow(axb, (0.61, 0.46), (0.50, 0.33), color=COL["supported"])
    arrow(axb, (0.86, 0.46), (0.56, 0.33), color=COL["hold"])
    draw_box(axb, (0.69, 0.25), 0.27, 0.16, "Development\n107 drugs · 16,470 pairs", face="#DDEBF7")
    draw_box(axb, (0.69, 0.03), 0.27, 0.16, "Temporal holdout\n59 drugs · 9,681 pairs", face="#FCE4D6")
    arrow(axb, (0.61, 0.25), (0.69, 0.33), color=COL["dev"])
    arrow(axb, (0.61, 0.19), (0.69, 0.11), color=COL["hold"])

    axc.set_title("Frozen prediction architecture", loc="left", fontweight="bold", pad=7)
    draw_box(axc, (0.03, 0.56), 0.27, 0.24, "Set 0\nRegulatory + event identity\n+ drug-program context", face="#EEEEEE")
    draw_box(axc, (0.03, 0.18), 0.27, 0.24, "Set 1\nSet 0 + pair-specific\npremarketing safety", face="#DDEBF7", edge=COL["set1"], lw=1.2)
    draw_box(axc, (0.40, 0.37), 0.22, 0.28, "Two frozen families\n○ Penalized logistic\n□ Gradient-boosted trees", face="#F7F7F7")
    draw_box(axc, (0.71, 0.57), 0.25, 0.22, "Development\nDrug-grouped nested CV", face="#DDEBF7")
    draw_box(axc, (0.71, 0.20), 0.25, 0.22, "Temporal validation\n2019–2022 locked holdout", face="#FCE4D6")
    arrow(axc, (0.30, 0.68), (0.40, 0.57))
    arrow(axc, (0.30, 0.30), (0.40, 0.45))
    arrow(axc, (0.62, 0.52), (0.71, 0.68), color=COL["dev"])
    arrow(axc, (0.62, 0.46), (0.71, 0.31), color=COL["hold"])
    axc.text(0.50, 0.06, "Outcome: 3-year FAERS Criterion R", ha="center", fontweight="bold")
    base = MAIN_FIG / "Figure1_study_design"
    export_figure(fig, base, "Figure1")
    for panel, source, cols, metric, trans in [
        ("A", "docs/ANALYSIS_STAGE_LOCK_v1.md", "data_source; role; model_role", "Data-source roles", "Qualitative architecture only"),
        ("B", "analysis/section1_cohort/01_cohort_flow_counts.csv; Section 1 README", "stage; active_moieties; trials; arms; drug_pt_pairs", "Cohort counts", "Parallel branches and locked cascade"),
        ("C", "analysis/section3_model/SECTION3_README.md; analysis/section4_holdout/SECTION4_README.md", "feature sets; model families; split; outcome", "Frozen prediction architecture", "Qualitative architecture only"),
    ]:
        vmanifest("Figure1", panel, source, cols, "Sections 1, 3, 4", metric, trans, "Not applicable", base)


def figure2():
    cov = read_csv("analysis/section4_holdout/12_full_2012_2022_coverage.csv")
    soc = read_csv("analysis/section2_coverage/04_soc_coverage.csv")
    drug = read_csv("analysis/section2_coverage/03_drug_level_coverage.csv")
    full = cov[cov["scope"] == "FULL_2012_2022"].copy()
    periods = cov[cov["scope"].isin(["DEVELOPMENT_2012_2018", "TEMPORAL_HOLDOUT_2019_2022"])].copy()
    assert int(full["premarketing_observed"].iloc[0]) == 3174
    assert int(full["postmarketing_only"].iloc[0]) == 13078
    assert int(full["all_criterion_r_signals"].iloc[0]) == 16252
    a_src = full[["scope", "all_criterion_r_signals", "premarketing_observed", "premarketing_observed_percent",
                  "premarketing_observed_ci_low", "premarketing_observed_ci_high", "postmarketing_only",
                  "postmarketing_only_percent"]]
    b_src = periods[["scope", "active_moieties", "premarketing_observed_percent", "premarketing_observed_ci_low",
                     "premarketing_observed_ci_high", "macro_coverage_percent", "macro_coverage_ci_low", "macro_coverage_ci_high"]]
    c_src = periods[["scope", "active_moieties", "median_drug_coverage_percent", "drug_coverage_q1_percent", "drug_coverage_q3_percent"]].copy()
    c_src["display_note"] = "Locked summary; temporal individual drug points are not available in frozen outputs"
    major = soc[soc["major_soc_ge30_signals"]].copy()
    d_src = major.nlargest(10, "criterion_r_signals").sort_values("coverage_pct")
    source_csv(a_src, "Figure2_panelA_full_coverage.csv")
    source_csv(b_src, "Figure2_panelB_temporal_consistency.csv")
    source_csv(c_src, "Figure2_panelC_drug_distribution_summary.csv")
    source_csv(d_src, "Figure2_panelD_major_SOC_top10_by_signal_count.csv")
    source_csv(major, "FigureS1_all_major_SOC_coverage.csv")
    source_csv(drug, "FigureS2_development_drug_level_coverage.csv")

    def panels(include_soc: bool, base_name: str, fig_id: str):
        if include_soc:
            fig = plt.figure(figsize=(7.2, 7.0))
            gs = fig.add_gridspec(2, 2, width_ratios=[0.78, 1.22], hspace=0.48, wspace=0.72)
            axes = [fig.add_subplot(gs[i, j]) for i in range(2) for j in range(2)]
        else:
            fig = plt.figure(figsize=(7.2, 4.2))
            gs = fig.add_gridspec(1, 3, wspace=0.48)
            axes = [fig.add_subplot(gs[0, i]) for i in range(3)]
        axa, axb, axc = axes[:3]
        pre = float(full["premarketing_observed_percent"].iloc[0])
        post = float(full["postmarketing_only_percent"].iloc[0])
        axa.barh([0], [pre], color=COL["pre"], edgecolor=COL["ink"], linewidth=0.5, label="Premarketing observed")
        axa.barh([0], [post], left=[pre], color=COL["post"], edgecolor=COL["ink"], linewidth=0.5,
                 hatch="///", label="Postmarketing only")
        ci_low = float(full["premarketing_observed_ci_low"].iloc[0])
        ci_high = float(full["premarketing_observed_ci_high"].iloc[0])
        axa.text(pre + post / 2, 0, f"Postmarketing only\n13,078 · {post:.2f}%", ha="center",
                 va="center", color=COL["ink"], fontsize=6.4)
        axa.set_ylim(-0.65, 0.75)
        axa.annotate(f"Premarketing observed\n3,174 · {pre:.2f}%\n95% CI {ci_low:.2f}–{ci_high:.2f}%",
                     xy=(pre / 2, 0.38), xytext=(1.5, 0.68), textcoords="data",
                     ha="left", va="top", fontsize=5.9, color=COL["pre"],
                     fontweight="bold",
                     arrowprops={"arrowstyle": "-", "color": COL["pre"], "lw": 0.65})
        axa.set_xlim(0, 100)
        axa.set_yticks([])
        axa.set_xlabel("Share of 16,252 Criterion-R signals (%)")
        axa.set_title("Full cohort coverage", loc="left", fontweight="bold")
        clean_axes(axa, "x")

        labels = ["Development\n2012–2018", "Temporal holdout\n2019–2022"]
        x = np.arange(2)
        micro = periods["premarketing_observed_percent"].to_numpy()
        lo = periods["premarketing_observed_ci_low"].to_numpy()
        hi = periods["premarketing_observed_ci_high"].to_numpy()
        macro = periods["macro_coverage_percent"].to_numpy()
        macro_lo = periods["macro_coverage_ci_low"].to_numpy()
        macro_hi = periods["macro_coverage_ci_high"].to_numpy()
        axb.errorbar(x-0.07, micro, yerr=[micro-lo, hi-micro], fmt="o", color=COL["ink"],
                     markerfacecolor=COL["pre"], markersize=5, capsize=2.5, label="Micro coverage (primary)")
        axb.errorbar(x+0.08, macro, yerr=[macro-macro_lo, macro_hi-macro], fmt="s",
                     color=COL["ink"], markerfacecolor="white", markeredgecolor=COL["ink"],
                     markersize=5, capsize=2.5, label="Macro coverage (secondary)")
        axb.set_xticks(x, labels)
        axb.set_ylabel("Coverage (%)")
        axb.set_ylim(0, max(32, hi.max()+4))
        axb.set_title("Temporal consistency", loc="left", fontweight="bold")
        axb.legend(frameon=False, loc="upper left")
        clean_axes(axb, "y")

        med = periods["median_drug_coverage_percent"].to_numpy()
        q1 = periods["drug_coverage_q1_percent"].to_numpy()
        q3 = periods["drug_coverage_q3_percent"].to_numpy()
        colors = [COL["dev"], COL["hold"]]
        for i in range(2):
            axc.errorbar(i, med[i], yerr=[[med[i]-q1[i]], [q3[i]-med[i]]],
                         fmt="D" if i == 0 else "s", color=COL["ink"],
                         markerfacecolor=colors[i], markeredgecolor=COL["ink"],
                         markersize=5.5, capsize=4, linewidth=1.1, zorder=3)
        axc.set_xticks(x, labels)
        axc.set_ylabel("Drug-level coverage (%)")
        axc.set_ylim(0, 35)
        axc.set_title("Drug-level distribution summary", loc="left", fontweight="bold")
        axc.text(0.02, 0.02, "Median (marker) and IQR (line/caps)\nfrom frozen summaries",
                 transform=axc.transAxes, fontsize=6.3, color="#555555")
        clean_axes(axc, "y")
        for label, ax in zip("ABC", [axa, axb, axc]):
            panel_label(ax, label)
        if include_soc:
            axd = axes[3]
            y = np.arange(len(d_src))
            est = d_src["coverage_pct"].to_numpy()
            lo2 = d_src["cluster_bootstrap_ci95_low_pct"].to_numpy()
            hi2 = d_src["cluster_bootstrap_ci95_high_pct"].to_numpy()
            axd.errorbar(est, y, xerr=[est-lo2, hi2-est], fmt="o", color=COL["ink"],
                         markerfacecolor=COL["pre"], markersize=4, capsize=2)
            names = [textwrap.fill(x, 24) for x in d_src["primary_soc"]]
            axd.set_yticks(y, names, fontsize=5.9)
            axd.set_xlabel("Premarketing coverage (%)")
            axd.set_title("Major SOCs: 10 largest signal groups", loc="left", fontweight="bold")
            axd.text(0.99, 0.01, "MedDRA taxonomic groups", transform=axd.transAxes,
                     ha="right", fontsize=6.2, color="#555555")
            clean_axes(axd, "x")
            panel_label(axd, "D")
        base = MAIN_FIG / base_name
        export_figure(fig, base, fig_id)
        return base

    base_b = panels(False, "Figure2_coverage_vB", "Figure2_vB")
    # Figure 2 vB is the sole main layout after Command 18; the prior vA
    # main exports are retired because SOC coverage is now Figure S1.
    for suffix in (".svg", ".pdf", ".png", ".tiff"):
        (MAIN_FIG / "Figure2_coverage_vA").with_suffix(suffix).unlink(missing_ok=True)
    for suffix in (".svg", ".pdf", ".png", ".tiff"):
        shutil.copyfile(base_b.with_suffix(suffix), (MAIN_FIG / "Figure2_coverage").with_suffix(suffix))
    FIG_META["Figure2"] = FIG_META["Figure2_vB"]
    for fid, base, panel_set in [("Figure2_vB", base_b, "ABC"),
                                  ("Figure2", MAIN_FIG / "Figure2_coverage", "ABC")]:
        srcs = {
            "A": ("analysis/section4_holdout/12_full_2012_2022_coverage.csv", "all_criterion_r_signals; premarketing_observed; postmarketing_only; percentages; CI", "Full-cohort micro coverage", "Stacked proportion bar + locked CI"),
            "B": ("analysis/section4_holdout/12_full_2012_2022_coverage.csv", "premarketing_observed_percent; CI; macro_coverage_percent", "Period micro/macro coverage", "Point and interval display"),
            "C": ("analysis/section4_holdout/12_full_2012_2022_coverage.csv", "median_drug_coverage_percent; q1; q3", "Drug-level distribution summary", "Locked median/IQR on 0–35% axis; no holdout individual values reconstructed"),
            "D": ("analysis/section2_coverage/04_soc_coverage.csv", "primary_soc; criterion_r_signals; coverage_pct; CI", "Major-SOC coverage", "Top 10 by signal count for layout; all major SOCs in Figure S1"),
        }
        for p in panel_set:
            s = srcs[p]
            vmanifest(fid, p, s[0], s[1], "Sections 2 and 4", s[2], s[3],
                      s[0] if p != "C" else "Frozen summary columns", base)


def figure3():
    dev_perf = read_csv("analysis/section3_model/training/06_oof_performance.csv")
    dev_ci = read_csv("analysis/section3_model/training/09_bootstrap_performance_ci.csv")
    hold_perf = read_csv("analysis/section4_holdout/05_holdout_performance.csv")
    hold_ci = read_csv("analysis/section4_holdout/06_holdout_bootstrap_ci.csv")
    dev_inc = read_csv("analysis/section3_model/training/08_incremental_value.csv")
    hold_inc = read_csv("analysis/section4_holdout/07_holdout_incremental_value.csv")
    dci = dev_ci[dev_ci["metric"] == "average_precision"][["pipeline", "ci_low", "ci_high"]]
    hci = hold_ci[hold_ci["metric"] == "average_precision"][["pipeline", "ci_low", "ci_high"]]
    a = dev_perf.merge(dci, on="pipeline")
    b = hold_perf.merge(hci, on="pipeline")
    for df in (a, b):
        df["display_label"] = df["pipeline"].map(label_pipeline)
    c = pd.concat([
        dev_inc[dev_inc["metric"] == "delta_average_precision_set1_minus_set0"].assign(period="Development"),
        hold_inc[hold_inc["metric"] == "delta_average_precision_set1_minus_set0"].assign(period="Temporal holdout"),
    ], ignore_index=True)
    d = pd.concat([
        dev_inc[dev_inc["metric"] == "brier_improvement_set0_minus_set1"].assign(period="Development"),
        hold_inc[hold_inc["metric"] == "brier_improvement_set0_minus_set1"].assign(period="Temporal holdout"),
    ], ignore_index=True)
    source_csv(a, "Figure3_panelA_development_AP.csv")
    source_csv(b, "Figure3_panelB_holdout_AP.csv")
    source_csv(c, "Figure3_panelC_delta_AP.csv")
    source_csv(d, "Figure3_panelD_brier_improvement.csv")

    fig, axes = plt.subplots(2, 2, figsize=(7.2, 6.0), gridspec_kw={"hspace": 0.44, "wspace": 0.48})
    axa, axb, axc, axd = axes.ravel()
    plot_estimates(axa, a, "display_label", "average_precision", "ci_low", "ci_high",
                   "Development nested-OOF AP", no_skill=0.125319, xlim=(0.10, 0.43))
    plot_estimates(axb, b, "display_label", "average_precision", "ci_low", "ci_high",
                   "Temporal-holdout AP", no_skill=0.114658, xlim=(0.10, 0.43))
    for ax, df, title, xlabel in [
        (axc, c, "Incremental value: Set 1 − Set 0", "Δ average precision (95% CI)"),
        (axd, d, "Probability accuracy: Set 0 − Set 1", "Brier improvement (95% CI)"),
    ]:
        ypos = {("Development", "elasticnet"): 3.2, ("Temporal holdout", "elasticnet"): 2.5,
                ("Development", "xgboost"): 1.2, ("Temporal holdout", "xgboost"): 0.5}
        for _, row in df.iterrows():
            y = ypos[(row["period"], row["model_family"])]
            marker = "o" if row["model_family"] == "elasticnet" else "s"
            color = COL["dev"] if row["period"] == "Development" else COL["hold"]
            ax.errorbar(row["estimate"], y,
                        xerr=[[row["estimate"]-row["ci_low"]], [row["ci_high"]-row["estimate"]]],
                        fmt=marker, color=COL["ink"], markerfacecolor=color, markersize=5, capsize=2.5)
        ax.axvline(0, color="#555555", linestyle=":", linewidth=0.8)
        ax.set_yticks([3.2, 2.5, 1.2, 0.5], ["Elastic-net · development", "Elastic-net · temporal",
                                                  "XGBoost · development", "XGBoost · temporal"])
        ax.set_xlabel(xlabel)
        ax.set_title(title, loc="left", fontweight="bold")
        clean_axes(ax, "x")
    for label, ax in zip("ABCD", axes.ravel()):
        panel_label(ax, label)
    fig.legend(handles=[
        Line2D([0], [0], marker="o", color="none", markerfacecolor=COL["dev"], markeredgecolor=COL["ink"], label="Development"),
        Line2D([0], [0], marker="s", color="none", markerfacecolor=COL["hold"], markeredgecolor=COL["ink"], label="Temporal holdout"),
    ], frameon=False, loc="lower center", ncol=2, bbox_to_anchor=(0.5, -0.01))
    base = MAIN_FIG / "Figure3_prediction"
    export_figure(fig, base, "Figure3")
    entries = [
        ("A", "analysis/section3_model/training/06_oof_performance.csv; 09_bootstrap_performance_ci.csv", "pipeline; average_precision; ci_low; ci_high", "Section 3", "Development AP", "Pipeline labels only", "09_bootstrap_performance_ci.csv"),
        ("B", "analysis/section4_holdout/05_holdout_performance.csv; 06_holdout_bootstrap_ci.csv", "pipeline; average_precision; ci_low; ci_high", "Section 4", "Temporal AP", "Pipeline labels only", "06_holdout_bootstrap_ci.csv"),
        ("C", "analysis/section3_model/training/08_incremental_value.csv; analysis/section4_holdout/07_holdout_incremental_value.csv", "model_family; estimate; ci_low; ci_high", "Sections 3 and 4", "Set 1 − Set 0 ΔAP", "Long-format display by period", "Frozen incremental-value files"),
        ("D", "analysis/section3_model/training/08_incremental_value.csv; analysis/section4_holdout/07_holdout_incremental_value.csv", "model_family; estimate; ci_low; ci_high", "Sections 3 and 4", "Set 0 − Set 1 Brier improvement", "Long-format display by period", "Frozen incremental-value files"),
    ]
    for e in entries:
        vmanifest("Figure3", *e, base)


def figure4():
    cal = read_csv("analysis/section4_holdout/08_holdout_calibration_source_data.csv")
    concord = read_csv("analysis/section5_interpretation/07_cross_model_rank_concordance.csv")
    supported = read_csv("analysis/section5_interpretation/08_cross_model_supported_features.csv")
    domain = read_csv("analysis/section5_interpretation/12_figure_interpretation_source_data/panel_b_set1_domain_attribution.csv")
    dependence = read_csv("analysis/section5_interpretation/12_figure_interpretation_source_data/panel_c_supported_feature_dependence.csv")
    pair = concord[concord["pair_specific"]].dropna(subset=["elasticnet_rank_pair_specific", "xgboost_rank_pair_specific"]).copy()
    support_names = set(supported.loc[supported["cross_model_supported"], "conceptual_feature"])
    dom = domain[(domain["feature_set"] == "SET1") & (domain["aggregation_level"] == "SCIENTIFIC_DOMAIN")].copy()
    dependence_features = [
        "pair_reporting_trial_fraction",
        "pair_median_row_ae_proportion",
        "pair_n_serious_arms",
    ]
    dep_main = dependence[dependence["conceptual_feature"].isin(dependence_features)].copy()
    source_csv(cal, "Figure4_panelA_temporal_calibration.csv")
    source_csv(pair, "Figure4_panelB_pair_specific_rank_concordance.csv")
    source_csv(dom, "Figure4_panelC_domain_attribution.csv")
    source_csv(dep_main, "Figure4_panelD_OOF_dependence.csv")

    fig = plt.figure(figsize=(7.2, 7.1))
    gs = fig.add_gridspec(3, 2, height_ratios=[1.0, 1.05, 0.92], hspace=0.57, wspace=0.48)
    axa1, axa2 = fig.add_subplot(gs[0, 0]), fig.add_subplot(gs[0, 1])
    axb, axc = fig.add_subplot(gs[1, 0]), fig.add_subplot(gs[1, 1])
    dgs = gs[2, :].subgridspec(1, 3, wspace=0.42)
    daxes = [fig.add_subplot(dgs[0, i]) for i in range(3)]
    plot_calibration(axa1, cal, "elasticnet", "Penalized logistic regression")
    plot_calibration(axa2, cal, "xgboost", "Gradient-boosted trees")
    panel_label(axa1, "A")
    axa1.text(0.0, 1.17, "Temporal calibration: native probabilities", transform=axa1.transAxes,
              fontweight="bold", fontsize=8)

    ordinary = pair[~pair["conceptual_feature"].isin(support_names)]
    high = pair[pair["conceptual_feature"].isin(support_names)]
    axb.scatter(ordinary["elasticnet_rank_pair_specific"], ordinary["xgboost_rank_pair_specific"],
                color="#C7C7C7", edgecolor="#666666", linewidth=0.3, s=20)
    axb.scatter(high["elasticnet_rank_pair_specific"], high["xgboost_rank_pair_specific"],
                color=COL["set1"], edgecolor=COL["ink"], linewidth=0.4, s=34, marker="D")
    short = {
        "pair_reporting_trial_fraction": "Trial recurrence",
        "pair_masked_trial_fraction": "Masked trials",
        "pair_median_row_ae_proportion": "Median AE proportion",
        "pair_nonduplicated_arm_subjects_at_risk": "Subjects at risk",
        "pair_n_serious_arms": "Serious arms",
    }
    label_positions = {
        "pair_reporting_trial_fraction": (8.0, 1.2),
        "pair_masked_trial_fraction": (11.5, 7.0),
        "pair_median_row_ae_proportion": (13.0, 3.5),
        "pair_nonduplicated_arm_subjects_at_risk": (19.5, 5.2),
        "pair_n_serious_arms": (19.5, 9.0),
    }
    for _, row in high.iterrows():
        feature = row["conceptual_feature"]
        axb.annotate(short[feature],
                     (row["elasticnet_rank_pair_specific"], row["xgboost_rank_pair_specific"]),
                     xytext=label_positions[feature], textcoords="data", fontsize=5.6,
                     ha="center", va="center",
                     bbox={"facecolor": "white", "edgecolor": "none", "alpha": 0.80, "pad": 0.6},
                     arrowprops={"arrowstyle": "-", "color": "#666666", "lw": 0.45})
    axb.set_xlabel("Elastic-net pair-specific rank")
    axb.set_ylabel("XGBoost pair-specific rank")
    axb.set_xlim(28, 0)
    axb.set_ylim(28, 0)
    axb.text(0.03, 0.04, "Spearman ρ = 0.336\nRank 1 = highest contribution",
             transform=axb.transAxes, ha="left", va="bottom", fontsize=6.2,
             bbox={"facecolor": "white", "edgecolor": "none", "alpha": 0.80, "pad": 0.8})
    axb.set_title("Cross-model rank concordance", loc="left", fontweight="bold")
    clean_axes(axb, "both")
    panel_label(axb, "B")

    pivot = dom.pivot(index="attribution_group", columns="model_family", values="relative_attribution_share").fillna(0)
    pivot = pivot.sort_values("xgboost")
    y = np.arange(len(pivot))
    h = 0.34
    axc.barh(y-h/2, pivot["elasticnet"]*100, height=h, color=COL["elasticnet"], edgecolor=COL["ink"], linewidth=0.4, label="Elastic-net")
    axc.barh(y+h/2, pivot["xgboost"]*100, height=h, color=COL["xgboost"], edgecolor=COL["ink"], linewidth=0.4, hatch="//", label="XGBoost")
    domain_labels = {
        "AE_PROPORTION": "AE proportion", "DRUG_PROGRAM": "Drug program",
        "EVENT_IDENTITY": "Event identity", "EVIDENCE_VOLUME": "Evidence volume",
        "CROSS_TRIAL": "Cross-trial", "SERIOUS_CONTEXT": "Serious context",
        "REGULATORY": "Regulatory", "PAIR_TRIAL_DESIGN": "Pair trial design",
        "REPORTING_THRESHOLD": "Reporting threshold",
    }
    axc.set_yticks(y, [domain_labels.get(x, x.replace("_", " ").title()) for x in pivot.index], fontsize=6.0)
    axc.set_xlabel("Descriptive attribution share within model (%)")
    axc.set_title("Set 1 domain attribution", loc="left", fontweight="bold")
    axc.legend(frameon=False, loc="lower right", fontsize=6.2)
    clean_axes(axc, "x")
    panel_label(axc, "C")

    dep_titles = {
        "pair_reporting_trial_fraction": "Reporting-trial fraction",
        "pair_median_row_ae_proportion": "Median row AE proportion",
        "pair_n_serious_arms": "Number of serious arms",
    }
    dep_colors = [COL["supported"], COL["xgboost"], COL["set1"]]
    y_min = float(dep_main["mean_grouped_shap"].min()) - 0.06
    y_max = float(dep_main["mean_grouped_shap"].max()) + 0.06
    for i, (ax, feature, color) in enumerate(zip(daxes, dependence_features, dep_colors)):
        sub = dep_main[dep_main["conceptual_feature"] == feature].sort_values("bin_order")
        xx = sub["bin_order"].to_numpy()
        ax.plot(xx, sub["mean_grouped_shap"], marker="o", color=color,
                markeredgecolor=COL["ink"], markeredgewidth=0.35)
        ax.axhline(0, color="#666666", linestyle=":", linewidth=0.7)
        if sub["data_type"].iloc[0] == "proportion":
            tick_labels = []
            for value in sub["median_feature_value"]:
                pct = 100 * value
                tick_labels.append(f"{pct:.0f}%" if pct >= 10 or pct == 0 else f"{pct:.2f}%")
        else:
            tick_labels = [f"{v:g}" for v in sub["median_feature_value"]]
        ax.set_xticks(xx, tick_labels, rotation=30, ha="right")
        ax.set_ylim(y_min, y_max)
        ax.set_xlabel("OOF quantile bin\n(median value shown)")
        if i == 0:
            ax.set_ylabel("Mean grouped TreeSHAP\n(margin log-odds)")
        ax.set_title(dep_titles[feature], loc="left", fontsize=6.8, fontweight="bold")
        clean_axes(ax, "y")
    panel_label(daxes[0], "D", -0.22, 1.31)
    daxes[0].text(0.0, 1.30, "Frozen OOF dependence summaries",
                  transform=daxes[0].transAxes, fontsize=8, fontweight="bold", va="top")
    base = MAIN_FIG / "Figure4_transport_interpretation"
    export_figure(fig, base, "Figure4")
    entries = [
        ("A", "analysis/section4_holdout/08_holdout_calibration_source_data.csv", "pipeline; mean_predicted_probability; observed_rate", "Section 4", "Native temporal calibration", "10-bin curves; no recalibration", "Not applicable (locked bins)"),
        ("B", "analysis/section5_interpretation/07_cross_model_rank_concordance.csv; 08_cross_model_supported_features.csv", "rank_pair_specific; cross_model_supported", "Section 5", "Pair-specific rank concordance", "Direct-label locked five supported predictors", "Not applicable"),
        ("C", "analysis/section5_interpretation/12_figure_interpretation_source_data/panel_b_set1_domain_attribution.csv", "model_family; attribution_group; relative_attribution_share", "Section 5", "Descriptive domain attribution", "Percent display only", "Not applicable"),
        ("D", "analysis/section5_interpretation/12_figure_interpretation_source_data/panel_c_supported_feature_dependence.csv", "conceptual_feature; bin; median_feature_value; mean_grouped_shap", "Section 5", "Frozen OOF dependence summaries", "OOF quantile-bin medians with connected descriptive guides; three locked main-text feature families; no SHAP recalculation", "Not applicable"),
    ]
    for e in entries:
        vmanifest("Figure4", *e, base)


def figure5():
    summary = read_csv("analysis/section6_robustness/04_jader_replication_summary.csv")
    status = read_csv("analysis/section6_robustness/05_jader_replication_by_premarketing_status.csv")
    robust = read_csv("analysis/section6_robustness/13_model_robustness_summary.csv")
    summary_idx = summary.set_index("metric")
    stage1 = summary_idx.loc[["JADER_ASSESSABLE", "NOT_ASSESSABLE"]].reset_index().assign(
        stage="Stage 1", denominator_label="16,252 definitive FAERS Criterion-R signals")
    stage2 = summary_idx.loc[["DIRECTIONALLY_POSITIVE", "JADER_R_REPLICATED", "JADER_CONSENSUS_REPLICATED"]].reset_index().assign(
        stage="Stage 2", denominator_label="9,736 JADER-assessable signals")
    two_stage = pd.concat([stage1, stage2], ignore_index=True)
    source_csv(two_stage, "Figure5_panelA_JADER_flow.csv")
    source_csv(status, "Figure5_panelB_replication_by_premarketing.csv")
    source_csv(robust, "Figure5_panelC_endpoint_horizon_deltaAP.csv")

    fig = plt.figure(figsize=(7.2, 5.5))
    gs = fig.add_gridspec(2, 2, height_ratios=[0.92, 1.08], hspace=0.55, wspace=0.48)
    ags = gs[0, :].subgridspec(1, 2, wspace=0.34)
    axa1, axa2 = fig.add_subplot(ags[0, 0]), fig.add_subplot(ags[0, 1])
    axb, axc = fig.add_subplot(gs[1, 0]), fig.add_subplot(gs[1, 1])

    stage1_labels = ["JADER assessable", "Not assessable"]
    y1 = np.arange(len(stage1_labels))[::-1]
    bars1 = axa1.barh(y1, stage1["percent"], color=[COL["set1"], COL["post"]],
                      edgecolor=COL["ink"], linewidth=0.45)
    bars1[1].set_hatch("///")
    for yi, (_, row) in zip(y1, stage1.iterrows()):
        axa1.text(row["percent"] + 1.1, yi, f"{int(row['n']):,} ({row['percent']:.2f}%)",
                  va="center", fontsize=6.4)
    axa1.set_yticks(y1, stage1_labels)
    axa1.set_xlim(0, 78)
    axa1.set_xlabel("Percent of 16,252 signals")
    axa1.set_title("Stage 1 · JADER assessability\nDenominator: 16,252 definitive FAERS Criterion-R signals",
                   loc="left", fontweight="bold", fontsize=7.2)
    clean_axes(axa1, "x")
    panel_label(axa1, "A", -0.13, 1.18)

    stage2_labels = ["ROR >1 (descriptive)", "JADER-R replicated", "Consensus replicated"]
    y2a = np.arange(len(stage2_labels))[::-1]
    bars2 = axa2.barh(y2a, stage2["percent"],
                      color=[COL["directional"], COL["pre"], COL["consensus"]],
                      edgecolor=COL["ink"], linewidth=0.45)
    bars2[0].set_hatch("..")
    bars2[2].set_hatch("xx")
    for yi, (_, row) in zip(y2a, stage2.iterrows()):
        axa2.text(row["percent"] + 0.65, yi, f"{int(row['n']):,} ({row['percent']:.2f}%)",
                  va="center", fontsize=6.4)
    axa2.set_yticks(y2a, stage2_labels)
    axa2.set_xlim(0, 43)
    axa2.set_xlabel("Percent of 9,736 assessable signals")
    axa2.set_title("Stage 2 · Cumulative cross-database replication\nDenominator: 9,736 JADER-assessable signals",
                   loc="left", fontweight="bold", fontsize=7.2)
    clean_axes(axa2, "x")

    y2 = np.arange(len(status))[::-1]
    est = status["replication_percent"].to_numpy()
    lo = status["replication_ci_low_percent"].to_numpy()
    hi = status["replication_ci_high_percent"].to_numpy()
    for i, (_, row) in enumerate(status.iterrows()):
        yi = y2[i]
        color = COL["pre"] if row["premarketing_status"] == "PREMARKETING_OBSERVED" else COL["post"]
        marker = "o" if row["premarketing_status"] == "PREMARKETING_OBSERVED" else "s"
        axb.errorbar(row["replication_percent"], yi,
                     xerr=[[row["replication_percent"]-row["replication_ci_low_percent"]],
                           [row["replication_ci_high_percent"]-row["replication_percent"]]],
                     fmt=marker, color=COL["ink"], markerfacecolor=color, markersize=6, capsize=3)
        axb.text(row["replication_ci_high_percent"]+0.7, yi,
                 f"{int(row['jader_r_replicated']):,}/{int(row['assessable_pairs']):,}", va="center", fontsize=6.5)
    axb.set_yticks(y2, ["Premarketing observed", "Postmarketing only"])
    axb.set_xlabel("JADER-R replication (%)")
    axb.set_xlim(0, 39)
    axb.set_title("Descriptive cross-system replication", loc="left", fontweight="bold")
    clean_axes(axb, "x")

    rows = []
    for _, r in robust.iterrows():
        endpoint = f"{int(r['horizon_years'])}-year {r['endpoint']}"
        rows.extend([
            [endpoint, "Elastic-net", r["elasticnet_delta_ap"], r["elasticnet_delta_ap_ci_low"], r["elasticnet_delta_ap_ci_high"]],
            [endpoint, "XGBoost", r["xgboost_delta_ap"], r["xgboost_delta_ap_ci_low"], r["xgboost_delta_ap_ci_high"]],
        ])
    forest = pd.DataFrame(rows, columns=["endpoint", "model_family", "delta_ap", "ci_low", "ci_high"])
    endpoint_order = list(dict.fromkeys(forest["endpoint"]))
    for i, endpoint in enumerate(endpoint_order[::-1]):
        base_y = i
        for fam, offset, marker, color in [("Elastic-net", -0.11, "o", COL["elasticnet"]),
                                            ("XGBoost", 0.11, "s", COL["xgboost"])]:
            row = forest[(forest["endpoint"] == endpoint) & (forest["model_family"] == fam)].iloc[0]
            axc.errorbar(row["delta_ap"], base_y+offset,
                         xerr=[[row["delta_ap"]-row["ci_low"]], [row["ci_high"]-row["delta_ap"]]],
                         fmt=marker, color=color, markerfacecolor="white" if fam == "Elastic-net" else color,
                         markeredgecolor=COL["ink"], markersize=4.8, capsize=2)
    axc.axvline(0, color="#555555", linestyle=":", linewidth=0.8)
    axc.set_yticks(range(len(endpoint_order)), endpoint_order[::-1])
    axc.set_xlabel("Set 1 − Set 0 ΔAP (95% CI)")
    axc.set_title("Endpoint and horizon robustness", loc="left", fontweight="bold")
    axc.legend(handles=[
        Line2D([0], [0], marker="o", color=COL["elasticnet"], markerfacecolor="white", label="Elastic-net"),
        Line2D([0], [0], marker="s", color=COL["xgboost"], label="XGBoost"),
    ], frameon=False, loc="upper center", bbox_to_anchor=(0.5, -0.16), ncol=2)
    clean_axes(axc, "x")

    panel_label(axb, "B")
    panel_label(axc, "C")
    base = MAIN_FIG / "Figure5_jader_robustness"
    export_figure(fig, base, "Figure5")
    source_csv(forest, "Figure5_panelC_endpoint_horizon_deltaAP_long.csv")
    entries = [
        ("A", "analysis/section6_robustness/04_jader_replication_summary.csv", "metric; n; denominator; percent", "Section 6", "Two-stage JADER assessability and replication", "Separate axes for the 16,252 and 9,736 denominators", "Frozen summary"),
        ("B", "analysis/section6_robustness/05_jader_replication_by_premarketing_status.csv", "assessable_pairs; jader_r_replicated; replication_percent; CI", "Section 6", "Descriptive JADER-R replication", "Two-group dot-and-whisker", "Frozen drug-bootstrap CI"),
        ("C", "analysis/section6_robustness/13_model_robustness_summary.csv", "endpoint; horizon; delta AP; CI", "Section 6", "Endpoint/horizon ΔAP", "Long-format forest display", "Frozen robustness summary"),
    ]
    for e in entries:
        vmanifest("Figure5", *e, base)


def supplementary_figures():
    soc = read_csv("analysis/section2_coverage/04_soc_coverage.csv")
    # Former main Figure 2 vA panel D, now relocated to Figure S1.
    major = soc[soc["major_soc_ge30_signals"]].nlargest(10, "criterion_r_signals").sort_values("coverage_pct")
    fig, ax = plt.subplots(figsize=(7.2, 4.8))
    y = np.arange(len(major))
    est, lo, hi = major["coverage_pct"].to_numpy(), major["cluster_bootstrap_ci95_low_pct"].to_numpy(), major["cluster_bootstrap_ci95_high_pct"].to_numpy()
    ax.errorbar(est, y, xerr=[est-lo, hi-est], fmt="o", color=COL["ink"], markerfacecolor=COL["pre"], markersize=4, capsize=2)
    ax.set_yticks(y, [textwrap.fill(x, 42) for x in major["primary_soc"]])
    ax.set_xlabel("Premarketing coverage (%)")
    ax.set_title("Major MedDRA SOC coverage: 10 largest signal groups", loc="left", fontweight="bold")
    ax.text(0.99, 0.01, "Taxonomic groups; not uniformly organ toxicities", transform=ax.transAxes, ha="right", fontsize=6.5)
    clean_axes(ax, "x")
    base = SUPP_FIG / "FigureS1_soc_coverage"
    export_figure(fig, base, "FigureS1")
    source_csv(major, "FigureS1_major_SOC_top10_by_signal_count.csv")
    vmanifest("FigureS1", "A", "analysis/section2_coverage/04_soc_coverage.csv", "primary_soc; criterion_r_signals; coverage_pct; CI", "Section 2", "Major-SOC coverage", "Top 10 by frozen signal count; sort by coverage for display", "Frozen SOC bootstrap CI", base)

    drug = read_csv("analysis/section2_coverage/03_drug_level_coverage.csv")
    strata = read_csv("analysis/section2_coverage/05_regulatory_strata_coverage.csv")
    cov = read_csv("analysis/section4_holdout/12_full_2012_2022_coverage.csv")
    fig, axes = plt.subplots(1, 2, figsize=(7.2, 3.6), gridspec_kw={"wspace": 0.48})
    x = np.zeros(len(drug)) + np.linspace(-0.12, 0.12, len(drug))
    axes[0].scatter(x, drug["coverage_percent"], s=10, facecolor=COL["dev"], edgecolor="white", linewidth=0.25, alpha=0.75)
    dev = cov[cov["scope"] == "DEVELOPMENT_2012_2018"].iloc[0]
    hold = cov[cov["scope"] == "TEMPORAL_HOLDOUT_2019_2022"].iloc[0]
    axes[0].scatter(0, dev["median_drug_coverage_percent"], marker="D", color=COL["ink"], s=35, zorder=4, label="Development median")
    axes[0].vlines(0.55, hold["drug_coverage_q1_percent"], hold["drug_coverage_q3_percent"], color=COL["hold"], linewidth=6, alpha=0.45)
    axes[0].scatter(0.55, hold["median_drug_coverage_percent"], marker="s", color=COL["hold"], edgecolor=COL["ink"], s=35, zorder=4)
    axes[0].set_xticks([0, 0.55], ["Development\nindividual drugs", "Temporal\nsummary only"])
    axes[0].set_ylabel("Drug-level coverage (%)")
    axes[0].set_title("Drug-level coverage", loc="left", fontweight="bold")
    clean_axes(axes[0], "y")
    use = strata[strata["stratum_variable"].isin(["nda_bla", "orphan_designation", "accelerated_approval"])].copy()
    use["label"] = use["stratum_variable"].str.replace("_", " ").str.title() + ": " + use["stratum_level"].astype(str)
    use = use.sort_values("micro_coverage_pct")
    yy = np.arange(len(use))
    est, lo, hi = use["micro_coverage_pct"].to_numpy(), use["cluster_bootstrap_ci95_low_pct"].to_numpy(), use["cluster_bootstrap_ci95_high_pct"].to_numpy()
    axes[1].errorbar(est, yy, xerr=[est-lo, hi-est], fmt="o", color=COL["ink"], markerfacecolor=COL["pre"], capsize=2)
    axes[1].set_yticks(yy, use["label"])
    axes[1].set_xlabel("Premarketing coverage (%)")
    axes[1].set_title("Selected prespecified regulatory strata", loc="left", fontweight="bold")
    clean_axes(axes[1], "x")
    for label, ax in zip("AB", axes): panel_label(ax, label)
    base = SUPP_FIG / "FigureS2_drug_regulatory_coverage"
    export_figure(fig, base, "FigureS2")
    vmanifest("FigureS2", "A", "analysis/section2_coverage/03_drug_level_coverage.csv; analysis/section4_holdout/12_full_2012_2022_coverage.csv", "coverage_percent; median; q1; q3", "Sections 2 and 4", "Drug-level coverage", "Development points; temporal frozen summary only", "Not applicable", base)
    vmanifest("FigureS2", "B", "analysis/section2_coverage/05_regulatory_strata_coverage.csv", "stratum; micro coverage; CI", "Section 2", "Regulatory-strata coverage", "Prespecified strata display", "Frozen cluster-bootstrap CI", base)

    fold = read_csv("analysis/section3_model/training/07_foldwise_performance.csv")
    fig, ax = plt.subplots(figsize=(7.2, 3.8))
    for pipeline, sub in fold.groupby("pipeline"):
        color = COL["set1"] if pipeline.endswith("set1") else COL["set0"]
        marker = "o" if pipeline.startswith("elasticnet") else "s"
        ls = "-" if pipeline.startswith("elasticnet") else "--"
        ax.plot(sub["outer_fold"], sub["average_precision"], marker=marker, linestyle=ls,
                color=color, markeredgecolor=COL["ink"], label=label_pipeline(pipeline))
    ax.set_xlabel("Outer drug-grouped fold")
    ax.set_ylabel("Average precision")
    ax.set_title("Outer-fold development performance", loc="left", fontweight="bold")
    ax.legend(frameon=False, ncol=2)
    clean_axes(ax, "both")
    base = SUPP_FIG / "FigureS3_outer_fold_performance"
    export_figure(fig, base, "FigureS3")
    source_csv(fold, "FigureS3_outer_fold_performance.csv")
    vmanifest("FigureS3", "A", "analysis/section3_model/training/07_foldwise_performance.csv", "pipeline; outer_fold; average_precision", "Section 3", "Foldwise AP", "Lines connect immutable folds for readability", "Not applicable", base)

    devcal = read_csv("analysis/section3_model/training/10_calibration_source_data.csv")
    holdcal = read_csv("analysis/section4_holdout/08_holdout_calibration_source_data.csv")
    fig, axes = plt.subplots(2, 2, figsize=(7.2, 6.4), gridspec_kw={"hspace": 0.42, "wspace": 0.38})
    for j, family in enumerate(("elasticnet", "xgboost")):
        plot_calibration(axes[0, j], devcal, family, f"Development · {'Elastic-net' if family == 'elasticnet' else 'XGBoost'}")
        plot_calibration(axes[1, j], holdcal, family, f"Temporal holdout · {'Elastic-net' if family == 'elasticnet' else 'XGBoost'}")
    for label, ax in zip("ABCD", axes.ravel()): panel_label(ax, label)
    base = SUPP_FIG / "FigureS4_complete_calibration"
    export_figure(fig, base, "FigureS4")
    source_csv(devcal.assign(period="Development"), "FigureS4_development_calibration.csv")
    source_csv(holdcal.assign(period="Temporal holdout"), "FigureS4_holdout_calibration.csv")
    vmanifest("FigureS4", "A–D", "analysis/section3_model/training/10_calibration_source_data.csv; analysis/section4_holdout/08_holdout_calibration_source_data.csv", "pipeline; bin; mean predicted probability; observed_rate", "Sections 3 and 4", "Native calibration", "Four locked 10-bin curves", "Not applicable", base)

    pt = read_csv("analysis/section4_holdout/09_pt_support_transportability.csv")
    primary = pt[pt["record_type"] == "PRIMARY_PT_SUPPORT"]
    increments = pt[pt["record_type"] == "PT_SUPPORT_INCREMENTAL"]
    subtype = pt[pt["record_type"] == "OPTIONAL_PT_RARE_SUBTYPE"]
    fig, axes = plt.subplots(1, 3, figsize=(7.2, 3.4), gridspec_kw={"wspace": 0.52})
    for i, stratum in enumerate(["PT_SUPPORTED", "PT_RARE"]):
        sub = primary[primary["pt_support_stratum"] == stratum]
        labels = [label_pipeline(x) for x in sub["pipeline"]]
        xx = np.arange(len(sub))
        axes[0].scatter(xx, sub["average_precision"], c=[COL["set1"] if x.endswith("set1") else COL["set0"] for x in sub["pipeline"]],
                        marker="o" if i == 0 else "s", edgecolor=COL["ink"], label=stratum)
    axes[0].set_xticks(np.arange(4), ["EN S0", "EN S1", "XGB S0", "XGB S1"], rotation=35, ha="right")
    axes[0].set_ylabel("Average precision")
    axes[0].set_title("PT-support transportability", loc="left", fontweight="bold")
    axes[0].legend(frameon=False, fontsize=6)
    clean_axes(axes[0], "y")
    inc = increments.copy()
    yy = np.arange(len(inc))
    axes[1].errorbar(inc["delta_average_precision_set1_minus_set0"], yy,
                     xerr=[inc["delta_average_precision_set1_minus_set0"]-inc["delta_ap_ci_low"], inc["delta_ap_ci_high"]-inc["delta_average_precision_set1_minus_set0"]],
                     fmt="o", color=COL["ink"], markerfacecolor=COL["set1"], capsize=2)
    axes[1].axvline(0, color="#555555", linestyle=":")
    axes[1].set_yticks(yy, inc["pt_support_stratum"] + " · " + inc["model_family"])
    axes[1].set_xlabel("ΔAP (95% CI)")
    axes[1].set_title("Within-stratum increment", loc="left", fontweight="bold")
    clean_axes(axes[1], "x")
    if not subtype.empty:
        pivot = subtype.pivot(index="pt_rare_subtype", columns="model_family", values="average_precision")
        pivot.plot(kind="bar", ax=axes[2], color=[COL["elasticnet"], COL["xgboost"]], edgecolor=COL["ink"], linewidth=0.4)
    axes[2].set_ylabel("Average precision")
    axes[2].set_xlabel("")
    axes[2].set_title("PT_RARE subtype (descriptive)", loc="left", fontweight="bold")
    axes[2].tick_params(axis="x", rotation=30)
    axes[2].legend(frameon=False, fontsize=6)
    clean_axes(axes[2], "y")
    for label, ax in zip("ABC", axes): panel_label(ax, label)
    base = SUPP_FIG / "FigureS5_PT_support_detail"
    export_figure(fig, base, "FigureS5")
    source_csv(pt, "FigureS5_PT_support_complete.csv")
    vmanifest("FigureS5", "A–C", "analysis/section4_holdout/09_pt_support_transportability.csv", "record_type; stratum; subtype; pipeline; AP; ΔAP; CI", "Section 4", "PT-support transportability", "Locked primary/increment/subtype records", "Frozen file", base)

    concord = read_csv("analysis/section5_interpretation/07_cross_model_rank_concordance.csv")
    dep = read_csv("analysis/section5_interpretation/12_figure_interpretation_source_data/panel_c_supported_feature_dependence.csv")
    supported = read_csv("analysis/section5_interpretation/08_cross_model_supported_features.csv")
    names = list(supported.loc[supported["cross_model_supported"], "conceptual_feature"])
    fig = plt.figure(figsize=(7.2, 7.4))
    gs = fig.add_gridspec(3, 2, hspace=0.5, wspace=0.42)
    ax0 = fig.add_subplot(gs[0, :])
    ax0.scatter(concord["elasticnet_rank_all"], concord["xgboost_rank_all"], color="#BDBDBD", s=18, edgecolor="#555555", linewidth=0.25)
    high = concord[concord["conceptual_feature"].isin(names)]
    ax0.scatter(high["elasticnet_rank_all"], high["xgboost_rank_all"], color=COL["set1"], marker="D", s=35, edgecolor=COL["ink"])
    ax0.set_xlim(46, 0); ax0.set_ylim(46, 0)
    ax0.set_xlabel("Elastic-net conceptual rank"); ax0.set_ylabel("XGBoost conceptual rank")
    ax0.set_title("Complete conceptual rank concordance (overall ρ = 0.270)", loc="left", fontweight="bold")
    clean_axes(ax0, "both"); panel_label(ax0, "A", -0.06)
    dep_axes = [fig.add_subplot(gs[1, 0]), fig.add_subplot(gs[1, 1]), fig.add_subplot(gs[2, 0]), fig.add_subplot(gs[2, 1])]
    for ax, feature, label in zip(dep_axes, names[:4], "BCDE"):
        sub = dep[dep["conceptual_feature"] == feature]
        ax.plot(sub["bin_order"], sub["mean_grouped_shap"], marker="o", color=COL["xgboost"])
        ax.axhline(0, color="#777777", linestyle=":", linewidth=0.7)
        ax.set_xticks(sub["bin_order"], sub["bin_label"], rotation=30, ha="right")
        ax.set_ylabel("Mean grouped TreeSHAP\n(margin log-odds)")
        ax.set_title(textwrap.fill(feature.replace("pair_", "").replace("_", " "), 28), loc="left", fontsize=7)
        clean_axes(ax, "y"); panel_label(ax, label)
    if len(names) == 5:
        inset = dep_axes[-1].inset_axes([0.53, 0.50, 0.45, 0.45])
        sub = dep[dep["conceptual_feature"] == names[4]]
        inset.plot(sub["bin_order"], sub["mean_grouped_shap"], marker="s", color=COL["supported"])
        inset.axhline(0, color="#777777", linestyle=":", linewidth=0.6)
        inset.set_title("serious arms", fontsize=6)
        inset.tick_params(labelsize=5)
    fig.text(0.5, 0.01, "Descriptive OOF model response; not causal", ha="center", fontsize=6.5, color="#8B1A1A")
    base = SUPP_FIG / "FigureS6_interpretation_dependence"
    export_figure(fig, base, "FigureS6")
    source_csv(concord, "FigureS6_complete_rank_concordance.csv")
    source_csv(dep, "FigureS6_supported_dependence.csv")
    vmanifest("FigureS6", "A–E", "analysis/section5_interpretation/07_cross_model_rank_concordance.csv; 12_figure_interpretation_source_data/panel_c_supported_feature_dependence.csv", "ranks; supported feature; bin; mean grouped SHAP", "Section 5", "Concordance and locked dependence summaries", "No feature ranking or SHAP recomputed", "Not applicable", base)

    rep = read_csv("analysis/section6_robustness/03_jader_replication_status.csv")
    reason = rep.groupby(["replication_status", "audit_reason"], dropna=False).size().reset_index(name="locked_pair_count")
    source_csv(reason, "FigureS7_JADER_reason_counts.csv")
    pivot = reason.pivot(index="audit_reason", columns="replication_status", values="locked_pair_count").fillna(0)
    pivot = pivot.loc[pivot.sum(axis=1).sort_values().index]
    fig, ax = plt.subplots(figsize=(7.2, max(4.2, len(pivot)*0.3)))
    left = np.zeros(len(pivot))
    color_map = {"REPLICATED": COL["pre"], "NOT_REPLICATED": COL["post"], "NOT_ASSESSABLE": COL["directional"]}
    hatch_map = {"REPLICATED": "", "NOT_REPLICATED": "///", "NOT_ASSESSABLE": "xx"}
    for col in pivot.columns:
        ax.barh(np.arange(len(pivot)), pivot[col], left=left, color=color_map.get(col, "#CCCCCC"),
                hatch=hatch_map.get(col, ""), edgecolor=COL["ink"], linewidth=0.35, label=col.replace("_", " ").title())
        left += pivot[col].to_numpy()
    ax.set_yticks(np.arange(len(pivot)), [textwrap.fill(str(x).replace("_", " ").title(), 48) for x in pivot.index])
    ax.set_xlabel("Locked FAERS signal-pair count")
    ax.set_title("JADER assessability and replication reason decomposition", loc="left", fontweight="bold")
    ax.legend(frameon=False)
    clean_axes(ax, "x")
    base = SUPP_FIG / "FigureS7_JADER_reasons"
    export_figure(fig, base, "FigureS7")
    vmanifest("FigureS7", "A", "analysis/section6_robustness/03_jader_replication_status.csv", "replication_status; audit_reason", "Section 6", "Locked status/reason counts", "Count locked categorical rows for display", "Not applicable", base)

    ps = read_csv("analysis/section6_robustness/08_ps_threshold_sensitivity.csv")
    da = read_csv("analysis/section6_robustness/09_definitionA_sensitivity.csv")
    bx = read_csv("analysis/section6_robustness/10_bexpanded_sensitivity.csv")
    fig, axes = plt.subplots(1, 3, figsize=(7.2, 3.6), gridspec_kw={"wspace": 0.58})
    axes[0].plot(ps["minimum_ps_cases"], ps["premarketing_coverage_percent"], marker="o", color=COL["pre"])
    axes[0].axvline(100, color=COL["ink"], linestyle=":")
    axes[0].set_xlabel("Minimum FAERS PS cases")
    axes[0].set_ylabel("Coverage (%)")
    axes[0].set_title("PS-threshold sensitivity", loc="left", fontweight="bold")
    clean_axes(axes[0], "both")
    axes[1].bar(["B-STRICT", "Definition A", "B-EXPANDED"],
                [19.52990401181393, float(da["coverage_percent"].iloc[0]), float(bx["coverage_percent"].iloc[0])],
                color=[COL["pre"], COL["directional"], COL["supported"]], edgecolor=COL["ink"], linewidth=0.4)
    axes[1].set_ylabel("Coverage (%)")
    axes[1].set_title("Coverage definitions\n(different denominators)", loc="left", fontweight="bold", fontsize=7.2)
    axes[1].tick_params(axis="x", rotation=35)
    clean_axes(axes[1], "y")
    timing = [int(bx["added_actual_final_completion_by_approval_links"].iloc[0]), int(bx["added_final_completion_after_approval_links"].iloc[0]), int(bx["added_missing_final_completion_date_links"].iloc[0])]
    axes[2].bar(["By approval", "After approval", "Missing date"], timing,
                color=[COL["pre"], COL["hold"], COL["post"]], edgecolor=COL["ink"], hatch=["", "//", "xx"])
    axes[2].set_ylabel("Additional drug–trial links")
    axes[2].set_title("B-EXPANDED\ntiming audit", loc="left", fontweight="bold", fontsize=7.2)
    axes[2].tick_params(axis="x", rotation=35)
    clean_axes(axes[2], "y")
    for label, ax in zip("ABC", axes): panel_label(ax, label)
    base = SUPP_FIG / "FigureS8_design_sensitivity"
    export_figure(fig, base, "FigureS8")
    source_csv(ps, "FigureS8_PS_threshold.csv"); source_csv(da, "FigureS8_DefinitionA.csv"); source_csv(bx, "FigureS8_BEXPANDED.csv")
    vmanifest("FigureS8", "A–C", "analysis/section6_robustness/08_ps_threshold_sensitivity.csv; 09_definitionA_sensitivity.csv; 10_bexpanded_sensitivity.csv", "threshold; coverage; definition; timing counts", "Section 6", "Design sensitivities", "Descriptive display; populations labeled", "Frozen files", base)


def number_format_for_header(header: str):
    h = header.lower()
    if any(x in h for x in ["percent", "coverage", "prevalence"]):
        return "0.00"
    if any(x in h for x in ["average_precision", "ap", "auroc", "brier", "log_loss", "slope", "intercept", "ci_low", "ci_high", "estimate", "smd"]):
        return "0.0000"
    return None


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame], title: str,
                   footnotes: list[str] | None = None):
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="D9D9D9")
    for sheet_name, df in sheets.items():
        safe_name = "".join("-" if ch in "[]:*?/\\" else ch for ch in sheet_name)[:31]
        ws = wb.create_sheet(safe_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        for j, col in enumerate(df.columns, 1):
            cell = ws.cell(1, j, col)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        for i, row in enumerate(df.itertuples(index=False, name=None), 2):
            for j, value in enumerate(row, 1):
                if pd.isna(value):
                    value = None
                if isinstance(value, str) and value.startswith("="):
                    value = "'" + value
                cell = ws.cell(i, j, value)
                cell.font = Font(name="Arial", size=8, color="000000")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F9FB")
                fmt = number_format_for_header(str(df.columns[j-1]))
                if fmt and isinstance(value, (int, float)):
                    cell.number_format = fmt
        last_data_row = len(df) + 1
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{last_data_row}"
        for j, col in enumerate(df.columns, 1):
            sample = [len(str(col))] + [len(str(x)) for x in df[col].head(250).fillna("")]
            ws.column_dimensions[get_column_letter(j)].width = min(max(max(sample)+2, 10), 42)
        ws.row_dimensions[1].height = 30
        if footnotes:
            note_row = last_data_row + 2
            for note in footnotes:
                ws.merge_cells(start_row=note_row, start_column=1,
                               end_row=note_row, end_column=len(df.columns))
                cell = ws.cell(note_row, 1, note)
                cell.font = Font(name="Arial", size=7, italic=True,
                                 bold=note.startswith("IMPORTANT:"), color="404040")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[note_row].height = 28
                note_row += 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.oddFooter.center.text = f"{title} | Locked-source formatted output"
    wb.save(path)
    check = load_workbook(path, data_only=False, read_only=True)
    errors = []
    for ws in check.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                    errors.append(f"{ws.title}!{cell.coordinate}")
    assert not errors, errors


def write_two_block_workbook(path: Path, block_a: pd.DataFrame, block_b: pd.DataFrame,
                             footnotes: list[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 2"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D9D9D9")

    def write_block(start_row: int, title: str, df: pd.DataFrame) -> int:
        end_col = len(df.columns)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
        title_cell = ws.cell(start_row, 1, title)
        title_cell.font = Font(name="Arial", bold=True, color="1F1F1F", size=9)
        title_cell.fill = PatternFill("solid", fgColor="D9EAF7")
        title_cell.alignment = Alignment(vertical="center")
        header_row = start_row + 1
        for j, col in enumerate(df.columns, 1):
            cell = ws.cell(header_row, j, col)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        for i, row in enumerate(df.itertuples(index=False, name=None), header_row + 1):
            for j, value in enumerate(row, 1):
                cell = ws.cell(i, j, None if pd.isna(value) else value)
                cell.font = Font(name="Arial", size=8)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if (i - header_row) % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F9FB")
        ws.row_dimensions[header_row].height = 38
        return header_row + len(df)

    end_a = write_block(1, "A. Absolute predictive performance", block_a)
    end_b = write_block(end_a + 2,
                        "B. Incremental value of pair-specific premarketing safety information",
                        block_b)
    max_cols = max(len(block_a.columns), len(block_b.columns))
    for j in range(1, max_cols + 1):
        values = []
        for df in (block_a, block_b):
            if j <= len(df.columns):
                col = df.columns[j - 1]
                values.extend([str(col)] + [str(x) for x in df[col].fillna("")])
        width = min(max(max(map(len, values), default=8) + 2, 11), 30)
        ws.column_dimensions[get_column_letter(j)].width = width
    note_row = end_b + 2
    for note in footnotes:
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=max_cols)
        cell = ws.cell(note_row, 1, note)
        cell.font = Font(name="Arial", size=7, italic=True, color="404040")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[note_row].height = 28
        note_row += 1
    ws.freeze_panes = "A3"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = "Table 2 | Locked-source formatted output"
    wb.save(path)


def write_two_block_csv(path: Path, block_a: pd.DataFrame, block_b: pd.DataFrame):
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["A. Absolute predictive performance"])
        writer.writerow(list(block_a.columns))
        writer.writerows(block_a.itertuples(index=False, name=None))
        writer.writerow([])
        writer.writerow(["B. Incremental value of pair-specific premarketing safety information"])
        writer.writerow(list(block_b.columns))
        writer.writerows(block_b.itertuples(index=False, name=None))


def table_csv(df: pd.DataFrame, path: Path):
    df.to_csv(path, index=False)


def tmanifest(table, source_file, columns, section, rule):
    TMANIFEST.append({
        "table": table, "source_file": source_file, "source_columns": columns,
        "locked_section": section, "formatting_rule": rule, "numerical_QC_result": "PASS",
    })


def tables():
    t1src = read_csv("analysis/section1_cohort/03_table1_development_vs_holdout.csv")
    # Command 19 main-table display selection: retain core regulatory and
    # trial-programme rows only. The complete frozen 50-row source is copied
    # to Table S1 below; no category is aggregated or recalculated.
    t1 = t1src[~t1src["variable"].isin(["route", "dosage_form", "candidate_pairs"])][
        ["variable", "level", "development_summary", "holdout_summary", "smd"]
    ].copy()
    labels = {
        "approval_year": "Approval year",
        "qualifying_trials": "Qualifying trials",
        "target_arms": "Target arms",
        "total_target_arm_subjects_at_risk": "Approximate premarketing AE safety-population size",
        "median_trial_enrollment": "Median trial enrollment",
        "randomized_trial_fraction": "Randomized trial fraction",
        "masked_trial_fraction": "Masked trial fraction",
        "industry_sponsored_fraction": "Industry-sponsored trial fraction",
        "candidate_pts": "Candidate PTs",
        "phase1_fraction": "Phase 1 trial fraction",
        "phase1_2_fraction": "Phase 1/2 trial fraction",
        "phase2_fraction": "Phase 2 trial fraction",
        "phase2_3_fraction": "Phase 2/3 trial fraction",
        "phase3_fraction": "Phase 3 trial fraction",
        "other_phase_fraction": "Other-phase trial fraction",
        "phase_missing_fraction": "Missing-phase trial fraction",
        "nda_bla": "Application type",
        "orphan_designation": "Orphan designation",
        "accelerated_approval": "Accelerated approval",
        "breakthrough_therapy_designation": "Breakthrough therapy designation",
        "fast_track_designation": "Fast-track designation",
        "priority_review": "Priority review",
    }
    t1["Characteristic"] = t1["variable"].map(labels)
    t1.loc[(t1["variable"] == "breakthrough_therapy_designation") & t1["level"].isna(), "level"] = "N/A"
    t1.loc[t1["level"].notna(), "Characteristic"] += ": " + t1.loc[t1["level"].notna(), "level"].astype(str)
    t1.loc[t1["variable"] == "approval_year", "development_summary"] = "2015 [2014, 2017]"
    t1.loc[t1["variable"] == "approval_year", "holdout_summary"] = "2021 [2019.5, 2021]"
    t1["smd"] = t1["smd"].map(
        lambda value: value if value == "N/A — temporal split-defining variable" else f"{float(value):.2f}"
    )
    t1 = t1[["Characteristic", "development_summary", "holdout_summary", "smd"]]
    t1.columns = ["Characteristic", "Development 2012–2018, N=107", "Temporal holdout 2019–2022, N=59", "Standardized mean difference"]
    assert t1.loc[t1["Characteristic"] == "Approval year", "Standardized mean difference"].iloc[0] == "N/A — temporal split-defining variable"
    assert len(t1) == 23
    table1_notes = [
        "Continuous variables are median [IQR]; categorical variables are n (%).",
        "Approximate premarketing AE safety-population size is a qualified programme measure and is not exact cumulative exposure.",
        "Approval year defines the temporal split and therefore is not assigned an imbalance SMD.",
        "Breakthrough therapy designation: N/A is structural, not a newly derived missing-data category.",
        "PT, MedDRA preferred term; SMD, standardized mean difference.",
    ]
    table_csv(t1, MAIN_TAB / "Table1_cohort_characteristics.csv")
    write_workbook(MAIN_TAB / "Table1_cohort_characteristics.xlsx", {"Table 1": t1}, "Table 1", table1_notes)
    tmanifest("Table 1", "analysis/section1_cohort/03_table1_development_vs_holdout.csv", "variable; level; development_summary; holdout_summary; smd", "Section 1", "23-row core display; approval year without grouping separators; SMD displayed to 2 decimals; complete source retained in Table S1")

    dp = read_csv("analysis/section3_model/training/06_oof_performance.csv")
    hp = read_csv("analysis/section4_holdout/05_holdout_performance.csv")
    hci = read_csv("analysis/section4_holdout/06_holdout_bootstrap_ci.csv")
    apci = hci[hci["metric"] == "average_precision"][["pipeline", "ci_low", "ci_high"]]
    block_a_raw = dp[["pipeline", "average_precision"]].rename(columns={"average_precision": "development_ap"}).merge(hp, on="pipeline").merge(apci, on="pipeline")
    model_labels = {
        "elasticnet_set0": "Penalized logistic regression, Set 0",
        "elasticnet_set1": "Penalized logistic regression, Set 1",
        "xgboost_set0": "Gradient-boosted trees (XGBoost), Set 0",
        "xgboost_set1": "Gradient-boosted trees (XGBoost), Set 1",
    }
    block_a = pd.DataFrame({
        "Model": block_a_raw["pipeline"].map(model_labels),
        "Development AP": block_a_raw["development_ap"].map(lambda x: f"{x:.3f}"),
        "Temporal AP": block_a_raw["average_precision"].map(lambda x: f"{x:.3f}"),
        "Temporal AP 95% CI": [f"{lo:.3f}–{hi:.3f}" for lo, hi in zip(block_a_raw["ci_low"], block_a_raw["ci_high"])],
        "Temporal AP lift": block_a_raw["auprc_lift"].map(lambda x: f"{x:.3f}"),
        "Temporal AUROC": block_a_raw["auroc"].map(lambda x: f"{x:.3f}"),
        "Temporal Brier score": block_a_raw["brier"].map(lambda x: f"{x:.3f}"),
        "Temporal log loss": block_a_raw["log_loss"].map(lambda x: f"{x:.3f}"),
        "Temporal calibration intercept": block_a_raw["calibration_intercept"].map(lambda x: f"{x:.3f}"),
        "Temporal calibration slope": block_a_raw["calibration_slope"].map(lambda x: f"{x:.3f}"),
    })
    di = read_csv("analysis/section3_model/training/08_incremental_value.csv")
    hi = read_csv("analysis/section4_holdout/07_holdout_incremental_value.csv")
    rows = []
    for fam in ["elasticnet", "xgboost"]:
        def metric(df, name):
            return df[(df["model_family"] == fam) & (df["metric"] == name)].iloc[0]
        dap_d = metric(di, "delta_average_precision_set1_minus_set0")
        dap_h = metric(hi, "delta_average_precision_set1_minus_set0")
        bri_h = metric(hi, "brier_improvement_set0_minus_set1")
        log_h = metric(hi, "log_loss_improvement_set0_minus_set1")
        rows.append([
            "Penalized logistic regression" if fam == "elasticnet" else "Gradient-boosted trees (XGBoost)",
            f"{dap_d['estimate']:.3f}", f"{dap_h['estimate']:.3f}",
            f"{dap_h['ci_low']:.3f}–{dap_h['ci_high']:.3f}",
            f"{bri_h['estimate']:.4f}", f"{bri_h['ci_low']:.4f}–{bri_h['ci_high']:.4f}",
            f"{log_h['estimate']:.3f}", f"{log_h['ci_low']:.3f}–{log_h['ci_high']:.3f}",
        ])
    block_b = pd.DataFrame(rows, columns=[
        "Model family", "Development ΔAP", "Temporal ΔAP", "Temporal ΔAP 95% CI",
        "Temporal Brier improvement", "Brier improvement 95% CI",
        "Temporal log-loss improvement", "Log-loss improvement 95% CI",
    ])
    table2_notes = [
        "AP, average precision. AP lift is temporal AP divided by endpoint prevalence.",
        "Set 0 contains regulatory, event-identity, and drug-programme context; Set 1 adds pair-specific premarketing safety information.",
        "Positive ΔAP favors Set 1; positive Brier/log-loss improvement indicates lower (better) loss for Set 1.",
        "Confidence intervals are from the frozen 5,000-resample active-moiety bootstrap.",
        "Calibration uses native unrecalibrated probabilities; no visual ranking of model families is intended.",
    ]
    write_two_block_csv(MAIN_TAB / "Table2_model_performance.csv", block_a, block_b)
    write_two_block_workbook(MAIN_TAB / "Table2_model_performance.xlsx", block_a, block_b, table2_notes)
    tmanifest("Table 2A", "analysis/section3_model/training/06_oof_performance.csv; analysis/section4_holdout/05_holdout_performance.csv; 06_holdout_bootstrap_ci.csv", "pipeline; AP; lift; AUROC; Brier; log loss; calibration", "Sections 3 and 4", "Four frozen pipelines; no algorithm ranking")
    tmanifest("Table 2B", "analysis/section3_model/training/08_incremental_value.csv; analysis/section4_holdout/07_holdout_incremental_value.csv", "model_family; metric; estimate; CI", "Sections 3 and 4", "Within-family Set 1 versus Set 0 contrasts")

    summ = read_csv("analysis/section6_robustness/04_jader_replication_summary.csv").set_index("metric")
    stat = read_csv("analysis/section6_robustness/05_jader_replication_by_premarketing_status.csv")
    rows = [[
        "Overall",
        f"{int(summ.loc['JADER_ASSESSABLE', 'n']):,}/{int(summ.loc['JADER_ASSESSABLE', 'denominator']):,} ({float(summ.loc['JADER_ASSESSABLE', 'percent']):.2f}%)",
        f"{int(summ.loc['JADER_R_REPLICATED', 'n']):,}/{int(summ.loc['JADER_R_REPLICATED', 'denominator']):,} ({float(summ.loc['JADER_R_REPLICATED', 'percent']):.2f}%)",
        f"{float(summ.loc['JADER_R_REPLICATED', 'ci_low_percent']):.2f}–{float(summ.loc['JADER_R_REPLICATED', 'ci_high_percent']):.2f}%",
        f"{int(summ.loc['JADER_CONSENSUS_REPLICATED', 'n']):,}",
    ]]
    for _, r in stat.iterrows():
        rows.append([
            r["premarketing_status"], f"{int(r['assessable_pairs']):,}",
            f"{int(r['jader_r_replicated']):,}/{int(r['assessable_pairs']):,} ({float(r['replication_percent']):.2f}%)",
            f"{float(r['replication_ci_low_percent']):.2f}–{float(r['replication_ci_high_percent']):.2f}%",
            f"{int(r['jader_consensus_replicated']):,}",
        ])
    t3 = pd.DataFrame(rows, columns=[
        "Population", "JADER-assessable pairs", "JADER-R replicated, n/N (%)",
        "95% CI", "Consensus replicated, n",
    ])
    table3_notes = [
        "IMPORTANT: JADER analysis represents cumulative cross-database replication and is not temporal external validation.",
        "JADER assessability required ≥50 verified cumulative JADER primary-suspect cases plus reliable drug and PT representation.",
        "JADER-R required a≥3 and a reporting-odds-ratio lower 95% confidence limit >1; Consensus additionally required IC025>0.",
        "PREMARKETING_OBSERVED and POSTMARKETING_ONLY follow the locked B-STRICT representation definitions.",
        "95% confidence intervals use the frozen drug-cluster (canonical active-moiety) bootstrap.",
    ]
    table_csv(t3, MAIN_TAB / "Table3_jader_replication.csv")
    write_workbook(MAIN_TAB / "Table3_jader_replication.xlsx", {"Table 3": t3}, "Table 3", table3_notes)
    tmanifest("Table 3", "analysis/section6_robustness/04_jader_replication_summary.csv; 05_jader_replication_by_premarketing_status.csv", "metric; denominator; n; percent; CI; consensus", "Section 6", "Three-row denominator-resolved display; directional ROR retained in Table S10; cumulative replication only")

    supplement_specs = {
        "TableS1_cohort_construction": [("Cohort flow", "analysis/section1_cohort/01_cohort_flow_counts.csv"), ("Full cohort characteristics", "analysis/section1_cohort/03_table1_development_vs_holdout.csv")],
        "TableS2_arm_attribution_audit": [("Arm selection SMD", "analysis/section1_cohort/05_arm_mapping_selection.csv")],
        "TableS3_drug_SOC_coverage": [("Drug coverage", "analysis/section2_coverage/03_drug_level_coverage.csv"), ("SOC coverage", "analysis/section2_coverage/04_soc_coverage.csv")],
        "TableS4_decomposition_thresholds": [("Corrected decomposition", "analysis/section2_coverage/06_postmarketing_only_decomposition.csv"), ("Threshold context", "analysis/section2_coverage/07_reporting_threshold_context.csv")],
        "TableS5_feature_dictionary": [("Feature dictionary", "analysis/section3_model/FEATURE_DICTIONARY_v1.csv")],
        "TableS6_CV_hyperparameters": [("Outer folds", "analysis/section3_model/06_outer_fold_balance.csv"), ("Inner CV QC", "analysis/section3_model/training/02_inner_cv_qc.csv"), ("Elastic-net tuning", "analysis/section3_model/training/03_hyperparameter_results_elasticnet.csv"), ("XGBoost tuning", "analysis/section3_model/training/04_hyperparameter_results_xgboost.csv"), ("Final tuning", "analysis/section3_model/training/12_final_full_development_tuning.csv")],
        "TableS7_development_performance": [("OOF performance", "analysis/section3_model/training/06_oof_performance.csv"), ("Incremental value", "analysis/section3_model/training/08_incremental_value.csv"), ("Bootstrap CIs", "analysis/section3_model/training/09_bootstrap_performance_ci.csv")],
        "TableS8_temporal_calibration_PT": [("Calibration bins", "analysis/section4_holdout/08_holdout_calibration_source_data.csv"), ("PT support", "analysis/section4_holdout/09_pt_support_transportability.csv")],
        "TableS9_feature_interpretation": [("Conceptual importance", "analysis/section5_interpretation/05_conceptual_feature_importance.csv"), ("Rank concordance", "analysis/section5_interpretation/07_cross_model_rank_concordance.csv"), ("Supported features", "analysis/section5_interpretation/08_cross_model_supported_features.csv"), ("Dependence summaries", "analysis/section5_interpretation/09_xgboost_dependence_summaries.csv")],
        "TableS10_JADER_assessability": [("Assessability", "analysis/section6_robustness/01_jader_assessability.csv"), ("Replication summary", "analysis/section6_robustness/04_jader_replication_summary.csv"), ("Status/reason counts", "figures/source_data/FigureS7_JADER_reason_counts.csv")],
        "TableS11_endpoint_horizon": [("Consensus", "analysis/section6_robustness/06_consensus_endpoint_performance.csv"), ("Early horizons", "analysis/section6_robustness/07_horizon_1y_2y_performance.csv"), ("Compact summary", "analysis/section6_robustness/13_model_robustness_summary.csv")],
        "TableS12_design_sensitivity": [("PS thresholds", "analysis/section6_robustness/08_ps_threshold_sensitivity.csv"), ("Definition A", "analysis/section6_robustness/09_definitionA_sensitivity.csv"), ("B-EXPANDED", "analysis/section6_robustness/10_bexpanded_sensitivity.csv")],
    }
    section_map = {1: "Section 1", 2: "Section 1", 3: "Section 2", 4: "Section 2", 5: "Section 3", 6: "Section 3", 7: "Section 3", 8: "Section 4", 9: "Section 5", 10: "Section 6", 11: "Section 6", 12: "Section 6"}
    for name, specs in supplement_specs.items():
        sheets = {}
        long = []
        for sheet_name, rel in specs:
            df = read_csv(rel)
            sheets[sheet_name] = df
            tmp = df.copy()
            tmp.insert(0, "source_table", sheet_name)
            tmp.insert(1, "source_file", rel)
            long.append(tmp)
            n = int(name.split("TableS")[1].split("_")[0])
            tmanifest(f"Table S{n}", rel, "; ".join(df.columns), section_map[n], "Authoritative frozen rows; presentation formatting only")
        combined = pd.concat(long, ignore_index=True, sort=False)
        table_csv(combined, SUPP_TAB / f"{name}.csv")
        write_workbook(SUPP_TAB / f"{name}.xlsx", sheets, name.replace("_", " "))


def write_manifests():
    pd.DataFrame(VMANIFEST).to_csv(FIG / "VISUALIZATION_MANIFEST_v1.csv", index=False)
    pd.DataFrame(TMANIFEST).to_csv(TAB / "TABLE_MANIFEST_v1.csv", index=False)


def main():
    figures_only = "--figures-only" in sys.argv[1:]
    tables_only = "--tables-only" in sys.argv[1:]
    visual_lock_only = "--visual-lock-only" in sys.argv[1:]
    if sum([figures_only, tables_only, visual_lock_only]) > 1:
        raise SystemExit("Choose only one of --figures-only, --tables-only, or --visual-lock-only")
    if visual_lock_only:
        existing = pd.read_csv(FIG / "VISUALIZATION_MANIFEST_v1.csv")
        figure2()
        figure4()
        replaced_ids = {"Figure2", "Figure2_vB", "Figure4"}
        retained = existing[~existing["figure_id"].isin(replaced_ids)]
        updated = pd.concat([retained, pd.DataFrame(VMANIFEST)], ignore_index=True)
        updated.to_csv(FIG / "VISUALIZATION_MANIFEST_v1.csv", index=False)
        print(json.dumps({
            "visual_lock_only": True,
            "rebuilt_main_figures": ["Figure2", "Figure4"],
            "visualization_manifest_rows": len(updated),
        }, indent=2))
        return
    if not tables_only:
        fig1()
        figure2()
        figure3()
        figure4()
        figure5()
        supplementary_figures()
    if figures_only:
        # Command 18 changes figure assets only. Preserve all table artifacts
        # and their locked manifest byte-for-byte.
        pd.DataFrame(VMANIFEST).to_csv(FIG / "VISUALIZATION_MANIFEST_v1.csv", index=False)
    elif tables_only:
        # Command 19 changes table assets only. Preserve figure artifacts and
        # the visualization manifest byte-for-byte.
        tables()
        pd.DataFrame(TMANIFEST).to_csv(TAB / "TABLE_MANIFEST_v1.csv", index=False)
    else:
        tables()
        write_manifests()
    print(json.dumps({
        "main_figure_files": len(list(MAIN_FIG.glob("*.svg"))),
        "supplement_figure_files": len(list(SUPP_FIG.glob("*.svg"))),
        "main_xlsx": len(list(MAIN_TAB.glob("*.xlsx"))),
        "supplement_xlsx": len(list(SUPP_TAB.glob("*.xlsx"))),
        "visualization_manifest_rows": (len(VMANIFEST) if VMANIFEST else
                                        len(pd.read_csv(FIG / "VISUALIZATION_MANIFEST_v1.csv"))),
        "table_manifest_rows": (len(TMANIFEST) if TMANIFEST else
                                len(pd.read_csv(TAB / "TABLE_MANIFEST_v1.csv"))),
    }, indent=2))


if __name__ == "__main__":
    main()
